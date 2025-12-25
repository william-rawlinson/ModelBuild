from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel
from io import BytesIO
import hashlib
import logging
from typing import Optional, List, Any, Dict, Tuple
from difflib import get_close_matches

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

router = APIRouter()
logger = logging.getLogger(__name__)

CANON_HEADERS = [
    "datapoint description",
    "base case value",
    "distribution",
    "standard error",
]

ALIASES = {
    "datapoint description": {
        # canonical
        "datapoint description",
        # spacing / punctuation
        "data point description",
        "datapoint desc",
        "data point desc",
        "description",
        "parameter description",
        "parameter desc",
        # health-econ / modelling language
        "parameter",
        "model parameter",
        "variable",
        "input",
        "input parameter",
        "quantity",
        "measure",
        # Excel-ish
        "name",
        "label",
    },

    "base case value": {
        # canonical
        "base case value",
        # spacing / punctuation
        "basecase value",
        "base-case value",
        "base value",
        "base case",
        "basecase",
        # statistical language
        "mean",
        "expected value",
        "expected",
        "central value",
        "point estimate",
        "estimate",
        # Excel-ish
        "value",
        "input value",
        "default value",
        # PSA / modelling
        "deterministic value",
        "base estimate",
    },

    "distribution": {
        # canonical
        "distribution",
        # common abbreviations
        "dist",
        "pdf",
        # PSA language
        "probability distribution",
        "statistical distribution",
        "uncertainty distribution",
        # Excel-ish
        "distribution type",
        "dist type",
        "distribution name",
    },

    "standard error": {
        # canonical
        "standard error",
        # spacing / punctuation
        "standarderror",
        "std error",
        "std. error",
        "std-error",
        # abbreviations
        "se",
        "stderr",
        "uncertainty",
        "error"
    },
}



class DataSheetMeta(BaseModel):
    filename: str
    size_bytes: int
    sha256: str
    sheet_name: str
    rows_extracted: int


class DataPoint(BaseModel):
    description: str
    base_case_value: Optional[float] = None
    distribution: Optional[str] = None
    standard_error: Optional[float] = None


class UploadModelDataSheetResponse(BaseModel):
    datapoints: List[DataPoint]
    meta: DataSheetMeta
    warnings: List[str]


def _sha256_bytes(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


def _norm(x: Any) -> str:
    # normalise for matching; keep it simple and deterministic
    return ("" if x is None else str(x)).strip().lower()


def _to_float(x: Any) -> Optional[float]:
    if x is None:
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(",", "")
    try:
        return float(s)
    except ValueError:
        return None


def _build_header_debug_message(
    *,
    header_row_raw: List[Any],
    header_map: Dict[str, int],
    missing: List[str],
) -> str:
    """
    Creates a very informative error message when required headers are missing.
    """
    raw_headers = [("" if h is None else str(h)) for h in header_row_raw]
    norm_headers = [_norm(h) for h in header_row_raw]

    # What we recognized (canonical -> actual header string + position)
    recognized_lines: List[str] = []
    for canon in CANON_HEADERS:
        if canon in header_map:
            idx = header_map[canon]
            actual = raw_headers[idx] if idx < len(raw_headers) else ""
            recognized_lines.append(
                f"- {canon!r} matched column {get_column_letter(idx+1)} (index {idx}) header {actual!r}"
            )
    if not recognized_lines:
        recognized_lines.append("- (none)")

    # Near misses for missing headers
    near_miss_lines: List[str] = []
    # Use norm_headers for similarity; show raw suggestion too where possible
    for canon in missing:
        candidates = get_close_matches(canon, norm_headers, n=3, cutoff=0.6)
        if candidates:
            # Map candidates back to raw header strings for readability
            suggestions = []
            for cand in candidates:
                # first occurrence index
                j = next((i for i, nh in enumerate(norm_headers) if nh == cand), None)
                raw = raw_headers[j] if j is not None else cand
                suggestions.append(f"{raw!r} (norm={cand!r}, col {get_column_letter((j or 0)+1)})")
            near_miss_lines.append(f"- For missing {canon!r}, closest headers: " + "; ".join(suggestions))
        else:
            near_miss_lines.append(f"- For missing {canon!r}, no close matches found in header transition_set.")

    expected_pretty = [
        "Datapoint description",
        "Base case value",
        "Distribution",
        "Standard error",
    ]

    # Show exactly what the sheet contained (raw + normalised)
    header_listing_lines = []
    for i, (raw, normed) in enumerate(zip(raw_headers, norm_headers)):
        if raw.strip() == "" and normed == "":
            continue
        header_listing_lines.append(
            f"- {get_column_letter(i+1)} (index {i}): raw={raw!r} | норм={normed!r}"
        )
    if not header_listing_lines:
        header_listing_lines = ["- (header transition_set appears blank)"]

    msg = (
        "Could not find all required columns in the first transition_set of the first sheet.\n\n"
        f"Missing required headers (canonical): {', '.join(repr(m) for m in missing)}\n\n"
        "Expected headers (exact text recommended):\n"
        + "\n".join(f"- {h}" for h in expected_pretty)
        + "\n\nRecognized headers:\n"
        + "\n".join(recognized_lines)
        + "\n\nHeaders found in the sheet (first transition_set):\n"
        + "\n".join(header_listing_lines)
        + "\n\nPotential near-misses:\n"
        + "\n".join(near_miss_lines)
        + "\n\nTip: ensure the header transition_set is transition_set 1, and the column names match exactly "
          "(case/spacing differences can matter in v1)."
    )
    return msg


def _find_header_map(header_row: List[Any]) -> Dict[str, int]:
    hdrs = [_norm(h) for h in header_row]
    out: Dict[str, int] = {}
    for canon in CANON_HEADERS:
        aliases = ALIASES[canon]
        for idx, h in enumerate(hdrs):
            if h in aliases:
                out[canon] = idx
                break
    return out


def extract_datapoints_from_excel(content: bytes) -> Tuple[List[DataPoint], str, List[str]]:
    warnings: List[str] = []
    wb = load_workbook(BytesIO(content), data_only=True, read_only=True)

    if not wb.sheetnames:
        raise ValueError("Workbook has no sheets.")

    ws = wb[wb.sheetnames[0]]
    sheet_name = ws.title

    rows_iter = ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return [], sheet_name, ["Sheet is empty."]

    header_map = _find_header_map(header)
    missing = [h for h in CANON_HEADERS if h not in header_map]
    if missing:
        raise ValueError(
            _build_header_debug_message(
                header_row_raw=header,
                header_map=header_map,
                missing=missing,
            )
        )

    out: List[DataPoint] = []

    for row in rows_iter:
        def get(canon: str):
            j = header_map[canon]
            return row[j] if j < len(row) else None

        desc = get("datapoint description")
        base = get("base case value")
        dist = get("distribution")
        se = get("standard error")

        if desc is None and base is None and dist is None and se is None:
            continue

        desc_s = ("" if desc is None else str(desc)).strip()
        if desc_s == "":
            warnings.append("Skipped a transition_set with missing 'Datapoint description'.")
            continue

        out.append(
            DataPoint(
                description=desc_s,
                base_case_value=_to_float(base),
                distribution=(None if dist is None or str(dist).strip() == "" else str(dist).strip()),
                standard_error=_to_float(se),
            )
        )

    return out, sheet_name, warnings


@router.post("/data-sheet", response_model=UploadModelDataSheetResponse)
async def upload_model_data_sheet(file: UploadFile = File(...)):
    filename = file.filename or ""
    lower = filename.lower()

    if not (lower.endswith(".xlsx") or lower.endswith(".xlsm")):
        raise HTTPException(
            status_code=400,
            detail="Only .xlsx/.xlsm files are supported in v1 (openpyxl cannot read .xls).",
        )

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Empty file.")

    max_bytes = 10 * 1024 * 1024
    if len(content) > max_bytes:
        raise HTTPException(status_code=400, detail="File too large (max 10MB).")

    sha256 = _sha256_bytes(content)

    try:
        datapoints, sheet_name, warnings = extract_datapoints_from_excel(content)
    except Exception as e:
        logger.exception("Failed to parse model data sheet")
        raise HTTPException(status_code=422, detail=f"Failed to parse Excel: {type(e).__name__}: {e}")

    logger.info(
        "Model data sheet uploaded: filename=%s size_bytes=%d sha256=%s sheet=%s rows=%d",
        filename, len(content), sha256, sheet_name, len(datapoints)
    )

    return UploadModelDataSheetResponse(
        datapoints=datapoints,
        meta=DataSheetMeta(
            filename=filename,
            size_bytes=len(content),
            sha256=sha256,
            sheet_name=sheet_name,
            rows_extracted=len(datapoints),
        ),
        warnings=warnings,
    )
